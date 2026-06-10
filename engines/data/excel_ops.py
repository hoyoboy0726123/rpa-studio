# -*- coding: utf-8 -*-
"""excel_ops — 純函式層:寫入 / 附加 / 依欄拆檔 / 規則拆檔 / 跨檔比對。

全部不依賴 ctx;actions.py 把這些包成 data.* / excel.* 步驟。
資料的通用型別是 list[dict](與 web.scrape_table 的輸出一致),
這樣「抓→填 Excel」可以直接把 scrape_table 寫進變數的值丟進來。

寫出的 xlsx 都做基本美化:粗體表頭、凍結首列、欄寬自動估算。
"""
from __future__ import annotations

import datetime as dt
import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# 共用:資料型別轉換 + 美化
# --------------------------------------------------------------------------- #
def to_records(data) -> list[dict]:
    """把輸入統一成 list[dict]。

    接受:list[dict] / DataFrame / 單一 dict(包成單列)。
    """
    if data is None:
        return []
    if isinstance(data, pd.DataFrame):
        return data.to_dict(orient="records")
    if isinstance(data, dict):
        return [data]
    if isinstance(data, (list, tuple)):
        out = []
        for r in data:
            if isinstance(r, dict):
                out.append(dict(r))
            else:
                out.append({"value": r})
        return out
    raise TypeError(f"to_records: 不支援的資料型別 {type(data)!r}")


def _union_columns(records: list[dict]) -> list[str]:
    """取所有列的欄位聯集,維持首見順序。"""
    cols: list[str] = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    return cols


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _beautify(ws, n_cols: int, n_rows: int) -> None:
    """套用表頭樣式、凍結首列、欄寬、框線。n_rows 不含表頭。"""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    # 欄寬:取該欄內容最長字元數(含表頭),夾在 8~60
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, n_rows + 2):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(8, min(60, max_len + 2))
    for r in range(1, n_rows + 2):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_sheet(ws, records: list[dict], columns: list[str] | None = None):
    """把 records 寫進一個 worksheet(含表頭),回 (n_cols, n_rows)。"""
    cols = columns or _union_columns(records)
    if not cols:
        cols = ["value"]
    ws.append(cols)
    for r in records:
        ws.append([_cell_value(r.get(c)) for c in cols])
    return len(cols), len(records)


def _cell_value(v):
    """openpyxl 不接受某些型別(list/dict)→ 轉字串;date 字串原樣。"""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool, dt.datetime, dt.date)):
        return v
    return str(v)


# --------------------------------------------------------------------------- #
# 寫入 / 附加
# --------------------------------------------------------------------------- #
def write_excel(data, path: str, sheet: str = "Sheet1",
                mode: str = "overwrite") -> dict:
    """把資料寫成 / 附加進 xlsx。

    mode = overwrite : 覆寫整張 sheet(其餘 sheet 保留)。
    mode = append    : 接在現有資料下方(欄位以現有表頭為準;檔不存在等同 overwrite)。
    回 {path, rows, cols, mode}。
    """
    from openpyxl import Workbook

    records = to_records(data)
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)

    if mode == "append" and os.path.exists(path):
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            # 以現有表頭為欄位順序
            existing_cols = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            existing_cols = [c for c in existing_cols if c is not None]
            if not existing_cols:
                existing_cols = _union_columns(records)
                ws.append(existing_cols)
            for r in records:
                ws.append([_cell_value(r.get(c)) for c in existing_cols])
            n_rows = ws.max_row - 1
            _beautify(ws, len(existing_cols), n_rows)
            wb.save(path)
            return {"path": path, "rows": len(records), "cols": len(existing_cols),
                    "mode": "append"}
        # sheet 不存在 → 新建一張
        ws = wb.create_sheet(title=sheet)
        n_cols, n_rows = _write_sheet(ws, records)
        _beautify(ws, n_cols, n_rows)
        wb.save(path)
        return {"path": path, "rows": n_rows, "cols": n_cols, "mode": "append"}

    # overwrite(或檔不存在)
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(title=sheet)
        # 確保新 sheet 不是被夾在最後造成困惑(非必要,留預設)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
    n_cols, n_rows = _write_sheet(ws, records)
    _beautify(ws, n_cols, n_rows)
    wb.save(path)
    return {"path": path, "rows": n_rows, "cols": n_cols, "mode": "overwrite"}


def csv_append(path: str, row: dict, include_header: bool = True) -> dict:
    """把一列(dict)附加到 CSV;檔不存在且 include_header 時先寫表頭。

    以 utf-8-sig 寫出(Excel 開中文不亂碼)。回 {path, wrote_header}。
    """
    import csv

    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)
    new_file = not os.path.exists(path) or os.path.getsize(path) == 0
    fieldnames = list(row.keys())
    with open(path, "a", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        wrote_header = False
        if new_file and include_header:
            writer.writeheader()
            wrote_header = True
        writer.writerow({k: ("" if v is None else v) for k, v in row.items()})
    return {"path": path, "wrote_header": wrote_header}


# --------------------------------------------------------------------------- #
# 依欄拆檔
# --------------------------------------------------------------------------- #
_SAFE_RE = re.compile(r'[\\/:*?"<>|]+')


def _safe_filename(name: str) -> str:
    """把分組值清成可當檔名的字串。"""
    s = _SAFE_RE.sub("_", str(name)).strip()
    s = re.sub(r"\s+", "_", s)
    return s or "未分類"


def split_by_column(input_path: str, group_col: str, out_dir: str,
                    prefix: str = "", sheet=None) -> list[dict]:
    """依 group_col 把資料拆成 N 個美化 xlsx,回索引 list[dict]。

    回每筆 {group, file, rows}。讀取走 robust_loader(容錯)。
    """
    from .robust_loader import read_table

    records = read_table(input_path, sheet=sheet)
    if records and group_col not in records[0]:
        # 容錯:欄名正規化後可能改變,嘗試比對正規化名稱
        from .robust_loader import normalize_header
        norm_target = normalize_header(group_col)
        keymap = {normalize_header(k): k for k in records[0].keys()}
        if norm_target in keymap:
            group_col = keymap[norm_target]
        else:
            raise KeyError(f"split_by_column: 找不到分組欄 {group_col!r};"
                           f"可用欄:{list(records[0].keys())}")

    os.makedirs(out_dir, exist_ok=True)
    groups: dict[str, list[dict]] = {}
    for r in records:
        key = r.get(group_col)
        key = "未分類" if key is None or str(key).strip() == "" else str(key)
        groups.setdefault(key, []).append(r)

    index: list[dict] = []
    for key, rows in groups.items():
        fname = f"{prefix}{_safe_filename(key)}.xlsx"
        fpath = os.path.join(out_dir, fname)
        write_excel(rows, fpath, sheet="Sheet1", mode="overwrite")
        index.append({"group": key, "file": fpath, "rows": len(rows)})

    # 索引檔
    index_path = os.path.join(out_dir, f"{prefix}_index.xlsx" if prefix else "_index.xlsx")
    write_excel(index, index_path, sheet="索引", mode="overwrite")
    return index


# --------------------------------------------------------------------------- #
# 規則拆檔
# --------------------------------------------------------------------------- #
def _match(value, match_type: str, pattern) -> bool:
    """單格是否符合規則。match_type: all/exact/prefix/suffix/contains/regex/range。"""
    mt = (match_type or "exact").lower()
    if mt == "all":
        return True
    sval = "" if value is None else str(value)
    if mt == "exact":
        return sval == str(pattern)
    if mt == "prefix":
        return sval.startswith(str(pattern))
    if mt == "suffix":
        return sval.endswith(str(pattern))
    if mt == "contains":
        return str(pattern) in sval
    if mt == "regex":
        try:
            return re.search(str(pattern), sval) is not None
        except re.error:
            return False
    if mt == "range":
        # pattern: [lo, hi] 或 "lo..hi";空界限視為無限
        lo = hi = None
        if isinstance(pattern, (list, tuple)) and len(pattern) == 2:
            lo, hi = pattern
        elif isinstance(pattern, str) and ".." in pattern:
            a, b = pattern.split("..", 1)
            lo = a.strip() or None
            hi = b.strip() or None
        try:
            x = float(sval)
        except (TypeError, ValueError):
            return False
        if lo is not None and x < float(lo):
            return False
        if hi is not None and x > float(hi):
            return False
        return True
    return False


def _render_template(template: str, *, category: str, data_yyyymm: str | None = None,
                     run_yyyymmdd: str | None = None) -> str:
    """套檔名模板,支援 {data_yyyymm}/{category}/{run_yyyymmdd}。"""
    now = dt.datetime.now()
    if data_yyyymm is None:
        first = now.replace(day=1)
        data_yyyymm = (first - dt.timedelta(days=1)).strftime("%Y%m")  # 預設上月
    if run_yyyymmdd is None:
        run_yyyymmdd = now.strftime("%Y%m%d")
    out = (template
           .replace("{data_yyyymm}", data_yyyymm)
           .replace("{category}", _safe_filename(category))
           .replace("{run_yyyymmdd}", run_yyyymmdd))
    if not out.lower().endswith(".xlsx"):
        out += ".xlsx"
    return out


def split_by_rules(input_path: str, rules: list[dict], out_dir: str,
                   filename_template: str = "{category}_{run_yyyymmdd}.xlsx",
                   sheet=None, data_yyyymm: str | None = None) -> list[dict]:
    """依規則拆檔。

    rules: [{name, column, match_type, pattern}]
      - name       : 該規則 / 分類名(對應模板 {category})
      - column     : 要比對的欄(match_type=all 時可省略)
      - match_type : all/exact/prefix/suffix/contains/regex/range
      - pattern    : 比對樣式(range 用 [lo,hi] 或 "lo..hi")
    一列符合多條規則時,會被寫進每一個符合的分類(非互斥)。
    回每筆 {name, file, rows}。
    """
    from .robust_loader import read_table

    records = read_table(input_path, sheet=sheet)
    os.makedirs(out_dir, exist_ok=True)

    buckets: dict[str, list[dict]] = {}
    order: list[str] = []
    for rule in rules:
        name = rule.get("name", "rule")
        col = rule.get("column")
        mt = rule.get("match_type", "exact")
        pat = rule.get("pattern")
        if name not in buckets:
            buckets[name] = []
            order.append(name)
        for r in records:
            val = r.get(col) if col else None
            if _match(val, mt, pat):
                buckets[name].append(r)

    out_index: list[dict] = []
    for name in order:
        rows = buckets[name]
        fname = _render_template(filename_template, category=name,
                                 data_yyyymm=data_yyyymm)
        fpath = os.path.join(out_dir, fname)
        write_excel(rows, fpath, sheet="Sheet1", mode="overwrite")
        out_index.append({"name": name, "file": fpath, "rows": len(rows)})
    return out_index


# --------------------------------------------------------------------------- #
# 跨檔比對
# --------------------------------------------------------------------------- #
_FILL_ADD = PatternFill("solid", fgColor="C6EFCE")   # 綠:新增
_FILL_DEL = PatternFill("solid", fgColor="FFC7CE")   # 紅:消失
_FILL_UP = PatternFill("solid", fgColor="BDD7EE")    # 藍:升
_FILL_DOWN = PatternFill("solid", fgColor="FCE4A6")  # 橙:降
_FILL_SAME = PatternFill("solid", fgColor="FFFFFF")  # 白:不變

_STATUS_LABEL = {
    "added": "新增", "removed": "消失",
    "up": "上升", "down": "下降", "same": "不變",
}


def _key_of(row: dict, key_cols: list[str]) -> tuple:
    return tuple(str(row.get(c, "")) for c in key_cols)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def diff_files(prev_path: str, curr_path: str, key_cols: list[str], value_col: str,
               out_path: str, threshold_pct: float = 0.0, sheet=None) -> dict:
    """兩檔比對,標 新增/消失/異動(±threshold%),產色標報告 xlsx。

    比對邏輯(以 key_cols 對齊):
      - 只在 curr     → 新增(綠)
      - 只在 prev     → 消失(紅)
      - 兩邊都有:看 value_col 變化百分比
          > +threshold% → 上升(藍)
          < -threshold% → 下降(橙)
          其餘           → 不變(白)
    回統計 {added, removed, up, down, same, out_path}。
    """
    from .robust_loader import read_table

    prev = read_table(prev_path, sheet=sheet)
    curr = read_table(curr_path, sheet=sheet)
    prev_map = {_key_of(r, key_cols): r for r in prev}
    curr_map = {_key_of(r, key_cols): r for r in curr}

    rows_out: list[dict] = []
    stats = {"added": 0, "removed": 0, "up": 0, "down": 0, "same": 0}

    all_keys = list(curr_map.keys()) + [k for k in prev_map if k not in curr_map]
    for k in all_keys:
        in_prev = k in prev_map
        in_curr = k in curr_map
        pv = _num(prev_map[k].get(value_col)) if in_prev else None
        cv = _num(curr_map[k].get(value_col)) if in_curr else None
        base = curr_map[k] if in_curr else prev_map[k]

        if in_curr and not in_prev:
            status, pct = "added", None
        elif in_prev and not in_curr:
            status, pct = "removed", None
        else:
            if pv is None or cv is None:
                status, pct = ("same", None)
            elif pv == 0:
                if cv == 0:
                    status, pct = "same", 0.0
                else:
                    status, pct = ("up" if cv > 0 else "down"), None
            else:
                pct = (cv - pv) / abs(pv) * 100.0
                if pct > threshold_pct:
                    status = "up"
                elif pct < -threshold_pct:
                    status = "down"
                else:
                    status = "same"
        stats[status] += 1
        rec = {c: base.get(c) for c in key_cols}
        rec["prev_" + value_col] = pv if in_prev else None
        rec["curr_" + value_col] = cv if in_curr else None
        rec["變化%"] = round(pct, 2) if pct is not None else None
        rec["狀態"] = _STATUS_LABEL[status]
        rec["_status"] = status
        rows_out.append(rec)

    # 寫色標報告
    _write_diff_report(rows_out, out_path, key_cols, value_col)
    stats["out_path"] = out_path
    return stats


def _write_diff_report(rows_out: list[dict], out_path: str,
                       key_cols: list[str], value_col: str) -> None:
    from openpyxl import Workbook

    d = os.path.dirname(out_path)
    if d:
        os.makedirs(d, exist_ok=True)
    cols = list(key_cols) + ["prev_" + value_col, "curr_" + value_col, "變化%", "狀態"]
    wb = Workbook()
    ws = wb.active
    ws.title = "比對報告"
    ws.append(cols)
    fill_by_status = {
        "added": _FILL_ADD, "removed": _FILL_DEL,
        "up": _FILL_UP, "down": _FILL_DOWN, "same": _FILL_SAME,
    }
    for r in rows_out:
        ws.append([_cell_value(r.get(c)) for c in cols])
        fill = fill_by_status.get(r.get("_status"), _FILL_SAME)
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).fill = fill
    _beautify(ws, len(cols), len(rows_out))
    # _beautify 會覆寫表頭樣式,但資料列填色保留(它只動表頭與框線)
    wb.save(out_path)

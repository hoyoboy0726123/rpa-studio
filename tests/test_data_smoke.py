# -*- coding: utf-8 -*-
"""data.* / excel.* 動作組 + robust_loader 冒煙測試。

不依賴 PySide6 / Playwright:直接組 ActionContext + run_flow,engine 設 None。
依賴 pandas / openpyxl(已在 requirements;缺則 pip install)。

執行(系統 python,專案根):
    python tests/test_data_smoke.py
全綠回 exit 0;任一失敗 raise AssertionError 並 exit 1。
"""
from __future__ import annotations

import os
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import Workbook, load_workbook

from core.registry import ActionContext, ACTIONS
from core.variables import VarStore
from core.schema import Flow, Step
from core.runner import run_flow
import engines.data  # noqa: F401  註冊 data.* / excel.*
from engines.data import robust_loader as rl


# --------------------------------------------------------------------------- #
class _FakeStore:
    def __init__(self):
        self.steps = []

    def log_step(self, run_id, step_id, action, status, ms=0, retries=0,
                 error="", screenshot=""):
        self.steps.append((step_id, action, status, error))


def _make_ctx(vars_init=None):
    return ActionContext(
        engine=None, vars=VarStore(vars_init or {}), vault=None,
        store=_FakeStore(), run_id="t", stop_event=None,
        log=lambda *_a, **_k: None, extra={},
    )


def _flow(steps):
    f = Flow(name="t", engine="web")
    f.steps = [Step.from_dict(s) for s in steps]
    return f


def _run(steps, vars_init=None):
    ctx = _make_ctx(vars_init)
    res = run_flow(_flow(steps), ctx)
    return ctx, res


# --------------------------------------------------------------------------- #
# robust_loader 純函式正規化
# --------------------------------------------------------------------------- #
def test_robust_normalizers():
    assert rl.normalize_number("1,234") == 1234
    assert rl.normalize_number("(1,234)") == -1234, rl.normalize_number("(1,234)")
    assert rl.normalize_number("$2,500.50") == 2500.5
    assert rl.normalize_number("12%") == 0.12
    assert rl.normalize_number("-87") == -87
    assert rl.normalize_number("abc") == "abc"

    assert rl.normalize_date("113/01/05") == "2024-01-05", rl.normalize_date("113/01/05")
    assert rl.normalize_date("民國113年1月5日") == "2024-01-05"
    assert rl.normalize_date("2024年3月") == "2024-03-01"
    assert rl.normalize_date("Jan 5, 2024") == "2024-01-05"
    assert rl.normalize_date("2024-12-31") == "2024-12-31"

    assert rl.normalize_header("  訂單（含稅）\n金額 ") == "訂單(含稅) 金額", \
        repr(rl.normalize_header("  訂單（含稅）\n金額 "))
    print("[OK] robust_loader normalizers(千分位/會計負數/貨幣/百分比/民國年/中文/英文月/欄名)")


def test_read_table_csv_dirty():
    """CSV:前置空白列 + 千分位 + 會計負數 + 民國年 + N/A → 讀出值正規化。"""
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "dirty.csv")
        # 前 2 列雜訊,第 3 列才是表頭
        content = (
            "月報表\n"
            ",,,\n"
            "料號,金額,異動日,備註\n"
            "A001,\"1,234\",113/01/05,N/A\n"
            "A002,\"(2,000)\",民國113年2月10日,正常\n"
        )
        with open(p, "w", encoding="utf-8-sig") as fh:
            fh.write(content)
        rows = rl.read_table(p)
        assert len(rows) == 2, rows
        assert rows[0]["料號"] == "A001"
        assert rows[0]["金額"] == 1234, rows[0]
        assert rows[1]["金額"] == -2000, rows[1]
        assert rows[0]["異動日"] == "2024-01-05", rows[0]
        assert rows[1]["異動日"] == "2024-02-10", rows[1]
        assert rows[0]["備註"] is None, rows[0]
    print("[OK] read_table CSV 前置空列+正規化")


def test_read_table_xlsx_dirty():
    """xlsx:前置空白列 + 千分位字串 → 偵測表頭、正規化。"""
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "dirty.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["季度報表"])          # 雜訊列
        ws.append([])                    # 空列
        ws.append(["產品", "數量", "日期"])
        ws.append(["X", "1,500", "113/03/01"])
        ws.append(["Y", "300", "2024年4月"])
        wb.save(p)
        rows = rl.read_table(p)
        assert len(rows) == 2, rows
        assert rows[0]["產品"] == "X"
        assert rows[0]["數量"] == 1500, rows[0]
        assert rows[0]["日期"] == "2024-03-01", rows[0]
        assert rows[1]["日期"] == "2024-04-01", rows[1]
    print("[OK] read_table xlsx 偵測表頭+正規化")


# --------------------------------------------------------------------------- #
# data.write_excel(抓→填 Excel:模擬 scrape_table 的 list[dict] 變數)
# --------------------------------------------------------------------------- #
def test_write_excel_overwrite_and_append():
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "report.xlsx")
        scraped = [
            {"序號": "SN001", "狀態": "通過"},
            {"序號": "SN002", "狀態": "失敗"},
        ]
        # overwrite:變數名 report_rows(模擬 web.scrape_table 的輸出)
        ctx, res = _run(
            [{"action": "data.write_excel",
              "params": {"var": "report_rows", "path": out, "sheet": "Sheet1",
                         "mode": "overwrite"}}],
            vars_init={"report_rows": scraped},
        )
        assert res.status == "completed", (res, ctx.store.steps)
        wb = load_workbook(out)
        ws = wb["Sheet1"]
        assert ws.max_row == 3, ws.max_row          # 表頭 + 2
        assert [c.value for c in ws[1]] == ["序號", "狀態"]

        # append:再追加 1 列
        ctx2, res2 = _run(
            [{"action": "data.write_excel",
              "params": {"var": "more", "path": out, "sheet": "Sheet1",
                         "mode": "append"}}],
            vars_init={"more": [{"序號": "SN003", "狀態": "通過"}]},
        )
        assert res2.status == "completed", (res2, ctx2.store.steps)
        wb2 = load_workbook(out)
        ws2 = wb2["Sheet1"]
        assert ws2.max_row == 4, ws2.max_row        # 表頭 + 3
        assert ws2.cell(row=4, column=1).value == "SN003"
    print("[OK] data.write_excel overwrite + append(scrape_table 變數 → Excel)")


def test_csv_append():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "log.csv")
        _run([{"action": "data.csv_append",
               "params": {"path": p, "row": {"a": 1, "b": "x"}, "include_header": True}}])
        _run([{"action": "data.csv_append",
               "params": {"path": p, "row": {"a": 2, "b": "y"}, "include_header": True}}])
        rows = rl.read_table(p)
        assert len(rows) == 2, rows
        assert rows[0]["a"] == 1 and rows[1]["b"] == "y", rows
    print("[OK] data.csv_append 追加多列")


# --------------------------------------------------------------------------- #
# excel.split(依欄拆檔)
# --------------------------------------------------------------------------- #
def test_excel_split():
    with tempfile.TemporaryDirectory() as d:
        src = os.path.join(d, "all.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["廠商", "金額"])
        for vendor, amt in [("A", 10), ("B", 20), ("A", 30), ("C", 5), ("B", 7)]:
            ws.append([vendor, amt])
        wb.save(src)

        out_dir = os.path.join(d, "out")
        ctx, res = _run([{"action": "excel.split",
                          "params": {"input_path": src, "group_col": "廠商",
                                     "out_dir": out_dir, "index_var": "idx"}}])
        assert res.status == "completed", (res, ctx.store.steps)
        idx = ctx.vars.get("idx")
        assert len(idx) == 3, idx                    # A/B/C 三檔
        files = {os.path.basename(e["file"]) for e in idx}
        assert {"A.xlsx", "B.xlsx", "C.xlsx"} <= files, files
        # A 檔應有 2 列資料
        a_file = [e for e in idx if e["group"] == "A"][0]
        assert a_file["rows"] == 2, a_file
        assert os.path.exists(os.path.join(out_dir, "_index.xlsx"))
    print("[OK] excel.split 依欄拆檔 + 索引")


# --------------------------------------------------------------------------- #
# excel.split_rules(規則拆檔 + 檔名模板)
# --------------------------------------------------------------------------- #
def test_excel_split_rules():
    with tempfile.TemporaryDirectory() as d:
        src = os.path.join(d, "items.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["料號", "金額"])
        for code, amt in [("SSD-1", 100), ("MEM-1", 200), ("SSD-2", 50),
                          ("HDD-1", 999), ("MEM-2", 30)]:
            ws.append([code, amt])
        wb.save(src)

        out_dir = os.path.join(d, "out")
        rules = [
            {"name": "固態硬碟", "column": "料號", "match_type": "prefix", "pattern": "SSD"},
            {"name": "記憶體", "column": "料號", "match_type": "contains", "pattern": "MEM"},
            {"name": "高額", "column": "金額", "match_type": "range", "pattern": [500, None]},
        ]
        ctx, res = _run([{"action": "excel.split_rules",
                          "params": {"input_path": src, "rules": rules,
                                     "out_dir": out_dir,
                                     "filename_template": "{category}_{data_yyyymm}.xlsx",
                                     "data_yyyymm": "202405",
                                     "index_var": "idx"}}])
        assert res.status == "completed", (res, ctx.store.steps)
        idx = {e["name"]: e for e in ctx.vars.get("idx")}
        assert idx["固態硬碟"]["rows"] == 2, idx
        assert idx["記憶體"]["rows"] == 2, idx
        assert idx["高額"]["rows"] == 1, idx
        # 檔名套模板
        assert os.path.basename(idx["固態硬碟"]["file"]) == "固態硬碟_202405.xlsx", idx
        assert os.path.exists(idx["記憶體"]["file"])
    print("[OK] excel.split_rules 規則拆檔 + 檔名模板")


# --------------------------------------------------------------------------- #
# excel.diff(兩檔比對 + 色標)
# --------------------------------------------------------------------------- #
def test_excel_diff():
    with tempfile.TemporaryDirectory() as d:
        prev = os.path.join(d, "prev.xlsx")
        curr = os.path.join(d, "curr.xlsx")
        out = os.path.join(d, "diff.xlsx")

        wb = Workbook(); ws = wb.active
        ws.append(["料號", "庫存"])
        for code, qty in [("A", 100), ("B", 50), ("C", 10), ("D", 80)]:
            ws.append([code, qty])
        wb.save(prev)

        wb = Workbook(); ws = wb.active
        ws.append(["料號", "庫存"])
        # A 不變;B 上升(50→70,+40%);C 下降(10→5,-50%);D 消失;E 新增
        for code, qty in [("A", 100), ("B", 70), ("C", 5), ("E", 33)]:
            ws.append([code, qty])
        wb.save(curr)

        ctx, res = _run([{"action": "excel.diff",
                          "params": {"prev_path": prev, "curr_path": curr,
                                     "key_cols": ["料號"], "value_col": "庫存",
                                     "out_path": out, "threshold_pct": 10,
                                     "result_var": "stats"}}])
        assert res.status == "completed", (res, ctx.store.steps)
        stats = ctx.vars.get("stats")
        assert stats["added"] == 1, stats          # E
        assert stats["removed"] == 1, stats        # D
        assert stats["up"] == 1, stats             # B
        assert stats["down"] == 1, stats           # C
        assert stats["same"] == 1, stats           # A
        # 報告可重開
        wb = load_workbook(out)
        ws = wb["比對報告"]
        assert ws.max_row == 6, ws.max_row         # 表頭 + 5
        # 標題列含「狀態」
        assert "狀態" in [c.value for c in ws[1]]
    print("[OK] excel.diff 新增/消失/上升/下降/不變 + 色標報告")


# --------------------------------------------------------------------------- #
# @action 註冊檢查
# --------------------------------------------------------------------------- #
def test_actions_registered():
    expected = {
        "data.read_table", "data.write_excel", "data.csv_append",
        "excel.split", "excel.split_rules", "excel.diff",
    }
    missing = expected - set(ACTIONS.keys())
    assert not missing, f"未註冊: {missing}"
    print(f"[OK] @action 註冊檢查:{sorted(expected)} 全在")


# --------------------------------------------------------------------------- #
def main():
    test_robust_normalizers()
    test_read_table_csv_dirty()
    test_read_table_xlsx_dirty()
    test_write_excel_overwrite_and_append()
    test_csv_append()
    test_excel_split()
    test_excel_split_rules()
    test_excel_diff()
    test_actions_registered()
    print("\nALL GREEN")


if __name__ == "__main__":
    main()
